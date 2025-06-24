from fastapi import FastAPI, UploadFile, HTTPException, File, Request, Query, Body, Form, Depends
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.middleware.gzip import GZipMiddleware
import os
import shutil
import uuid
import pdfplumber
import pandas as pd
import numpy as np
import cv2
import pytesseract
from PIL import Image
import io
from typing import List, Dict, Optional, Any
import xml.etree.ElementTree as ET
from datetime import datetime
import aiofiles
import json
from pydantic import BaseModel
import asyncio
from concurrent.futures import ProcessPoolExecutor
import logging
import sys
import traceback
import  uvicorn
import mimetypes
from fastapi import BackgroundTasks
from pathlib import Path
import re
from validation_utils import validate_table_data, get_validation_summary, validate_financial_data, ValidationError
from bank_statement_parser import process_bank_statement
from bank_matcher import BankMatcher
from gst_helper import GSTHelper
import fitz
from PyPDF2 import PdfReader
from audit_logger import audit_logger
import multiprocessing
from cachetools import TTLCache
import hashlib
from functools import lru_cache, wraps
import psutil
import threading
import time

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Performance optimization - file processing cache
file_cache = TTLCache(maxsize=100, ttl=3600)  # Cache for 1 hour

def get_file_hash(file_path: str) -> str:
    """Generate a hash for file content to use as cache key"""
    hash_sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()[:16]  # Use first 16 chars for shorter keys

def cached_file_processing(func):
    """Decorator for caching file processing results"""
    @wraps(func)
    def wrapper(file_path: str, *args, **kwargs):
        try:
            # Create cache key from file hash and function args
            file_hash = get_file_hash(file_path)
            cache_key = f"{func.__name__}_{file_hash}_{str(args)}_{str(sorted(kwargs.items()))}"
            
            # Check cache first
            if cache_key in file_cache:
                logger.info(f"Cache hit for {func.__name__}: {file_path}")
                return file_cache[cache_key]
            
            # Process file if not in cache
            result = func(file_path, *args, **kwargs)
            
            # Store in cache
            file_cache[cache_key] = result
            logger.info(f"Cached result for {func.__name__}: {file_path}")
            
            return result
        except Exception as e:
            logger.error(f"Error in cached processing: {str(e)}")
            # Fallback to direct function call if caching fails
            return func(file_path, *args, **kwargs)
    
    return wrapper

app = FastAPI(
    title="PDF Tally Converter API",
    description="API for converting PDF files",
    version="1.0.0"
)

# Add CORS middleware before any routes
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://localhost:5175",
        "http://localhost:3000"
    ],  # Frontend URLs
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add GZip compression
app.add_middleware(GZipMiddleware, minimum_size=1000)

@app.middleware("http")
async def log_requests(request: Request, call_next):
    # Get the start timestamp
    start_time = datetime.now()
    
    # Process the request
    response = await call_next(request)
    
    # Log the request after it's processed
    duration = (datetime.now() - start_time).total_seconds()
    
    # Only log API endpoints, skip static files
    if request.url.path.startswith("/api") or request.url.path in ["/upload", "/convert", "/validate"]:
        audit_logger.log_action(
            action_type="api_request",
            summary=f"{request.method} {request.url.path}",
            metadata={
                "duration": duration,
                "status_code": response.status_code,
                "client_host": request.client.host if request.client else None,
                "user_agent": request.headers.get("user-agent")
            }
        )
    
    return response

@app.get("/test")
async def test():
    """Test endpoint to verify the API is working"""
    return {"message": "API is working!"}

@app.get("/debug")
async def debug(request: Request):
    """Debug endpoint to show environment information"""
    return {
        "cwd": os.getcwd(),
        "files_in_cwd": os.listdir(),
        "python_path": sys.path,
        "env_vars": dict(os.environ),
        "request_url": str(request.url),
        "request_headers": dict(request.headers)
    }

@app.get("/")
async def root():
    """Root endpoint"""
    logger.info("Root endpoint accessed")
    return {
        "message": "PDF Tally Converter API",
        "status": "running",
        "timestamp": datetime.now().isoformat(),
        "available_endpoints": [
            "/",
            "/test",
            "/debug",
            "/health"
        ]
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    logger.info("Health check endpoint accessed")
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "environment": {
            "cwd": os.getcwd(),
            "python_version": sys.version
        }
    }

# Create necessary directories
UPLOAD_DIR = "uploads"
CONVERTED_DIR = "converted"
CHUNK_SIZE = 1024 * 1024  # 1MB chunks for file handling
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(CONVERTED_DIR, exist_ok=True)

# Initialize process pool for CPU-intensive tasks with optimal worker count
MAX_WORKERS = min(32, (multiprocessing.cpu_count() or 1) + 4)
process_pool = ProcessPoolExecutor(max_workers=MAX_WORKERS)

# Initialize bank matcher
bank_matcher = BankMatcher()

# Initialize GST helper
gst_helper = GSTHelper()

# Import the hybrid parser
try:
    from hybrid_bank_parser import parse_with_hybrid_engine, HybridBankParser
    HYBRID_PARSER_AVAILABLE = True
    logger.info("Hybrid bank parser loaded successfully")
except ImportError as e:
    HYBRID_PARSER_AVAILABLE = False
    logger.warning(f"Hybrid bank parser not available: {str(e)}")

class ConversionError(Exception):
    pass

class CellMetadata(BaseModel):
    error: Optional[bool] = None
    confidence: Optional[float] = None
    status: Optional[str] = None
    originalValue: Optional[Any] = None

class TableCell(BaseModel):
    value: Any
    metadata: CellMetadata

class TableRow(BaseModel):
    cells: Dict[str, TableCell]

    class Config:
        arbitrary_types_allowed = True

class TableData(BaseModel):
    headers: List[str]
    rows: List[TableRow]

    class Config:
        arbitrary_types_allowed = True

class EditHistory(BaseModel):
    timestamp: int
    rowId: str
    columnKey: str
    oldValue: Any
    newValue: Any

class SavePayload(BaseModel):
    fileId: str
    originalData: TableData
    modifiedData: TableData
    editHistory: List[EditHistory]

async def save_upload_file(upload_file: UploadFile) -> tuple[str, str]:
    """Save uploaded file and return the file path with optimized streaming"""
    file_id = str(uuid.uuid4())
    file_extension = os.path.splitext(upload_file.filename)[1]
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_extension}")
    
    try:
        # Use larger chunk size for better performance
        OPTIMIZED_CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
        
        async with aiofiles.open(file_path, 'wb') as out_file:
            while True:
                chunk = await upload_file.read(OPTIMIZED_CHUNK_SIZE)
                if not chunk:
                    break
                await out_file.write(chunk)
        
        return file_path, file_id
    except Exception as e:
        # Clean up the file if there's an error
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=str(e))

@cached_file_processing
def extract_tables_from_pdf(file_path: str, use_hybrid: bool = True) -> pd.DataFrame:
    """Extract tables from PDF using hybrid parser with fallback to existing parser"""
    logger.info(f"Extracting tables from PDF: {file_path} (hybrid={use_hybrid})")
    
    if not os.path.exists(file_path):
        logger.error(f"PDF file not found: {file_path}")
        return pd.DataFrame()
    
    # Try hybrid parser first if available and enabled
    if use_hybrid and HYBRID_PARSER_AVAILABLE:
        try:
            logger.info("Using hybrid bank parser")
            df = parse_with_hybrid_engine(file_path, enable_advanced=True)
            if df is not None and not df.empty:
                logger.info(f"Hybrid parser successful, extracted {len(df)} rows")
                return df
            else:
                logger.warning("Hybrid parser returned empty result, falling back to original")
        except Exception as e:
            logger.error(f"Hybrid parser failed: {str(e)}, falling back to original parser")
    
    # Fallback to original parser logic
    logger.info("Using original extraction logic")
    return _extract_tables_original(file_path)

def _extract_tables_original(file_path: str) -> pd.DataFrame:
    """Original table extraction logic"""
    try:
        # First try to validate PDF file
        try:
            with open(file_path, 'rb') as f:
                # Try PyMuPDF first
                try:
                    doc = fitz.open(file_path)
                    if doc.page_count == 0:
                        raise ValueError("PDF has no pages")
                    doc.close()
                except Exception as e:
                    logger.warning(f"PyMuPDF validation failed: {str(e)}")
                    # Try PyPDF2 as fallback
                    reader = PdfReader(f)
                    if len(reader.pages) == 0:
                        raise ValueError("PDF has no pages")
        except Exception as e:
            logger.error(f"PDF validation failed: {str(e)}")
            return pd.DataFrame()

        # Try general table extraction first
        logger.info("Starting pdfplumber table extraction...")
        try:
            with pdfplumber.open(file_path) as pdf:
                all_tables = []
                for page_num, page in enumerate(pdf.pages, 1):
                    logger.info(f"Processing page {page_num} of {len(pdf.pages)}")
                    
                    # Try different table finding strategies
                    table_settings = [
                        {"vertical_strategy": "text", "horizontal_strategy": "text"},
                        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                        {"vertical_strategy": "lines_strict", "horizontal_strategy": "lines_strict"},
                        {
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "text_tolerance": 3,
                            "text_x_tolerance": 3,
                            "intersection_tolerance": 3,
                            "snap_tolerance": 3,
                            "join_tolerance": 3,
                        }
                    ]
                    
                    tables = []
                    for settings in table_settings:
                        if not tables:  # Only try next strategy if no tables found
                            try:
                                found_tables = page.find_tables(settings)
                                if found_tables:
                                    tables = [table.extract() for table in found_tables]
                                    if tables and any(table for table in tables):
                                        logger.info(f"Found {len(tables)} tables with settings: {settings}")
                                        break
                            except Exception as e:
                                logger.warning(f"Table extraction failed with settings {settings}: {str(e)}")
                    
                    if tables:
                        # Filter out None tables and empty tables
                        valid_tables = [table for table in tables if table and len(table) > 1]
                        all_tables.extend(valid_tables)
                    else:
                        logger.warning(f"No tables found on page {page_num}")
                
                if all_tables:
                    logger.info(f"Found {len(all_tables)} tables total")
                    # Use the first table's headers
                    first_table = all_tables[0]
                    if not first_table or len(first_table) == 0:
                        logger.warning("First table is empty")
                        return pd.DataFrame()
                    
                    headers = first_table[0]
                    if not headers or not any(headers):
                        logger.warning("No valid headers found in first table")
                        return pd.DataFrame()
                    
                    # Combine all rows from all tables
                    rows = []
                    for table in all_tables:
                        if table and len(table) > 1:
                            rows.extend(table[1:])  # Skip header row for subsequent tables
                    
                    if not rows:
                        logger.warning("No data rows found in tables")
                        return pd.DataFrame()
                    
                    logger.info(f"Creating DataFrame with {len(headers)} columns and {len(rows)} rows")
                    df = pd.DataFrame(rows, columns=headers)
                    df = df.fillna('')  # Replace NaN with empty string
                    
                    # Clean up the data
                    df = df.replace(r'^\s*$', '', regex=True)  # Replace whitespace-only cells
                    df = df.replace(r'\s+', ' ', regex=True)   # Normalize whitespace
                    
                    # Try to detect if this is tabular data
                    if len(df.columns) >= 3 and len(df) > 0:
                        logger.info("Valid tabular data found, processing columns...")
                        # Standardize column names
                        column_mapping = {
                            col: col.lower().replace(' ', '_') if col else f'col_{i}'
                            for i, col in enumerate(df.columns)
                        }
                        df = df.rename(columns=column_mapping)
                        
                        # Ensure required columns exist
                        required_columns = {
                            'date': ['date', 'tran_date', 'transaction_date', 'value_date'],
                            'narration': ['narration', 'description', 'particulars', 'details'],
                            'amount': ['amount', 'debit', 'credit', 'dr', 'cr'],
                            'balance': ['balance', 'running_balance', 'closing_balance']
                        }
                        
                        # Map existing columns to standard names
                        for std_col, variants in required_columns.items():
                            if std_col not in df.columns:
                                # Find first matching variant
                                for var in variants:
                                    if var in df.columns:
                                        df = df.rename(columns={var: std_col})
                                        break
                                # If no variant found, add empty column
                                if std_col not in df.columns:
                                    df[std_col] = ''
                        
                        # Clean up date column
                        if 'date' in df.columns:
                            try:
                                df['date'] = pd.to_datetime(df['date'], format='mixed', errors='coerce')
                                df['date'] = df['date'].dt.strftime('%Y-%m-%d')
                            except Exception as e:
                                logger.warning(f"Error processing date column: {str(e)}")
                        
                        # Clean up amount columns
                        for col in ['amount', 'balance']:
                            if col in df.columns:
                                try:
                                    df[col] = df[col].apply(lambda x: str(x).replace('₹', '').replace(',', '').strip())
                                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                                except Exception as e:
                                    logger.warning(f"Error processing {col} column: {str(e)}")
                        
                        logger.info(f"Successfully processed tabular data with shape: {df.shape}")
                        return df
                    else:
                        logger.warning("Extracted data doesn't look like valid tabular data")
        except Exception as e:
            logger.error(f"pdfplumber table extraction failed: {str(e)}")
            logger.error(traceback.format_exc())
            
        # If general table extraction fails or doesn't look like tabular data,
        # try processing as bank statement
        logger.info("Trying bank statement parser as fallback...")
        try:
            df = process_bank_statement(file_path)
            if df is not None and not df.empty:
                logger.info("Successfully extracted data using bank statement parser")
                return df
            else:
                logger.warning("Bank statement parser returned empty DataFrame")
        except Exception as e:
            logger.error(f"Bank statement parser failed: {str(e)}")
            logger.error(traceback.format_exc())
        
        # If both methods fail, return empty DataFrame
        logger.warning("All extraction methods failed, returning empty DataFrame")
        return pd.DataFrame()
            
    except Exception as e:
        logger.error(f"Error in table extraction: {str(e)}")
        logger.error(traceback.format_exc())
        return pd.DataFrame()

@cached_file_processing
def process_image_ocr(file_path: str) -> pd.DataFrame:
    """Process image using OCR to extract tabular data with fallback when tesseract is not available"""
    logger.info(f"Processing image using OCR: {file_path}")
    
    try:
        # Read image
        img = cv2.imread(file_path)
        if img is None:
            logger.error(f"Failed to read image file: {file_path}")
            raise ConversionError("Failed to read image file")

        logger.info(f"Image loaded successfully with shape: {img.shape}")
        
        # First try using tesseract if available
        try:
            import pytesseract
            logger.info("Using Tesseract OCR")
            return _process_with_tesseract(img)
        except ImportError:
            logger.warning("Tesseract not available, using alternative OCR")
        except Exception as e:
            if "tesseract is not installed" in str(e).lower():
                logger.warning("Tesseract not installed, using alternative OCR")
            else:
                logger.error(f"Tesseract failed: {str(e)}, using alternative OCR")
        
        # Try hybrid parser
        if HYBRID_PARSER_AVAILABLE:
            try:
                logger.info("Using hybrid parser for image")
                from hybrid_bank_parser import HybridBankParser
                parser = HybridBankParser(enable_advanced_detection=True)
                
                # Extract cells directly from image
                from hybrid_bank_parser import TableRegion
                region = TableRegion(0, 0, img.shape[1], img.shape[0], 1.0, 0)
                cells = parser.extract_cells_from_region(img, region)
                
                if cells:
                    parsed_table = parser.reconstruct_table(cells)
                    parsed_table = parser.map_to_schema(parsed_table)
                    parsed_table = parser.clean_data(parsed_table)
                    
                    if parsed_table.rows:
                        df = pd.DataFrame(parsed_table.rows, columns=parsed_table.headers)
                        logger.info(f"Hybrid parser successful: {df.shape}")
                        return df
            except Exception as e:
                logger.warning(f"Hybrid parser failed: {str(e)}")
        
        # Use simple OCR fallback
        try:
            logger.info("Using simple OCR fallback")
            from simple_ocr_fallback import extract_with_simple_ocr
            df = extract_with_simple_ocr(img)
            if not df.empty:
                logger.info(f"Simple OCR successful: {df.shape}")
                return df
        except Exception as e:
            logger.warning(f"Simple OCR fallback failed: {str(e)}")
        
        # Final fallback: return structured dummy data to indicate we found the image
        logger.warning("All OCR methods failed, returning placeholder data")
        return pd.DataFrame({
            'date': [datetime.now().strftime('%Y-%m-%d')],
            'description': ['Image processed - OCR unavailable'],
            'amount': [0.0],
            'balance': [0.0]
        })
        
    except ConversionError:
        raise
    except Exception as e:
        logger.error(f"Error in OCR processing: {str(e)}")
        logger.error(traceback.format_exc())
        raise ConversionError(f"Failed to process image: {str(e)}")

def _process_with_tesseract(img: np.ndarray) -> pd.DataFrame:
    """Process image using Tesseract OCR"""
    import pytesseract
    
    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Improve image quality for OCR
    # Apply threshold to get better text recognition
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # OCR with different configurations
    custom_config = r'--oem 3 --psm 6'  # Treat the image as a single uniform block of text
    
    try:
        text = pytesseract.image_to_string(thresh, config=custom_config)
        logger.info(f"Tesseract OCR completed, extracted text length: {len(text)}")
    except Exception as e:
        logger.warning(f"Tesseract with custom config failed: {str(e)}, trying default config")
        text = pytesseract.image_to_string(thresh)
    
    if not text or not text.strip():
        logger.error("No text detected in image")
        return pd.DataFrame()
    
    # Convert OCR text to DataFrame
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if not lines:
        logger.error("No valid lines found in OCR text")
        return pd.DataFrame()
        
    logger.info(f"Found {len(lines)} lines of text")
    
    # Try to detect table structure
    # Look for lines that might be headers (containing common table keywords)
    header_keywords = ['date', 'amount', 'description', 'balance', 'debit', 'credit', 'transaction']
    header_line_idx = -1
    
    for i, line in enumerate(lines):
        line_lower = line.lower()
        if any(keyword in line_lower for keyword in header_keywords):
            # Check if this line has multiple "columns" (spaces between words)
            words = line.split()
            if len(words) >= 3:  # At least 3 columns
                header_line_idx = i
                logger.info(f"Found potential header line at index {i}: {line}")
                break
    
    if header_line_idx == -1:
        # If no clear header found, assume first line is header
        header_line_idx = 0
        logger.warning("No clear header line found, using first line as header")
    
    # Extract headers and data
    headers = lines[header_line_idx].split()
    data_lines = lines[header_line_idx + 1:]
    
    if not headers:
        logger.error("No headers found")
        return pd.DataFrame()
    
    if not data_lines:
        logger.error("No data lines found")
        return pd.DataFrame()
    
    logger.info(f"Headers: {headers}")
    logger.info(f"Data lines count: {len(data_lines)}")
    
    # Convert data lines to rows
    data = []
    for line in data_lines:
        if line.strip():
            # Split line into fields
            # Try to split by whitespace but handle cases where values might contain spaces
            words = line.split()
            if len(words) >= len(headers):
                # If we have more words than headers, try to group them intelligently
                if len(words) > len(headers):
                    # Keep first and last fields, combine middle ones
                    row = [words[0]]  # First field (likely date)
                    if len(headers) > 2:
                        # Combine middle fields (likely description)
                        middle_words = words[1:-len(headers)+2]
                        row.append(' '.join(middle_words))
                        # Add remaining fields
                        row.extend(words[-len(headers)+2:])
                    else:
                        row.extend(words[1:len(headers)])
                else:
                    row = words
                
                # Pad or trim row to match header count
                while len(row) < len(headers):
                    row.append('')
                row = row[:len(headers)]
                
                data.append(row)
            else:
                logger.warning(f"Skipping line with insufficient columns: {line}")
    
    if not data:
        logger.error("No valid data rows found")
        return pd.DataFrame()
    
    logger.info(f"Creating DataFrame with {len(headers)} columns and {len(data)} rows")
    df = pd.DataFrame(data, columns=headers)
    
    # Clean up column names
    df.columns = [col.lower().replace(' ', '_') for col in df.columns]
    
    # Basic data cleaning
    df = df.fillna('')
    df = df.replace(r'^\s*$', '', regex=True)
    
    logger.info(f"Successfully created DataFrame from Tesseract OCR with shape: {df.shape}")
    return df

def create_tally_xml(df: pd.DataFrame, output_path: str) -> None:
    """Create Tally XML from DataFrame"""
    logger.info(f"Creating Tally XML at: {output_path}")
    try:
        # Create the root element
        root = ET.Element("ENVELOPE")
        
        # Add header
        header = ET.SubElement(root, "HEADER")
        ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
        
        # Add body
        body = ET.SubElement(root, "BODY")
        importdata = ET.SubElement(body, "IMPORTDATA")
        requestdesc = ET.SubElement(importdata, "REQUESTDESC")
        reportname = ET.SubElement(requestdesc, "REPORTNAME").text = "Vouchers"
        
        # Add data
        requestdata = ET.SubElement(importdata, "REQUESTDATA")
        
        # Convert each row to XML
        for _, row in df.iterrows():
            voucher = ET.SubElement(requestdata, "VOUCHER")
            for col in df.columns:
                if pd.notna(row[col]):  # Only add non-null values
                    ET.SubElement(voucher, col.upper().replace(" ", "")).text = str(row[col])
        
        # Create the XML file
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ")
        tree.write(output_path, encoding="utf-8", xml_declaration=True)
        
        logger.info("Successfully created Tally XML file")
        
    except Exception as e:
        logger.error(f"Error creating Tally XML: {str(e)}")
        logger.error(traceback.format_exc())
        raise

def convert_file(file_path: str, file_id: str, output_formats: List[str] = ["xlsx", "csv", "xml"]) -> Dict:
    """Convert a file and return the extracted data"""
    try:
        # Extract data from file
        if file_path.lower().endswith('.pdf'):
            data = extract_tables_from_pdf(file_path)
        elif file_path.lower().endswith(('.png', '.jpg', '.jpeg')):
            data = process_image_ocr(file_path)
        else:
            raise ConversionError(f"Unsupported file format: {file_path}")

        # Get headers
        headers = data.columns.tolist()

        # Convert DataFrame to rows while preserving numeric types
        rows = []
        for _, row in data.iterrows():
            row_dict = {}
            for col in headers:
                value = row[col]
                # Convert numpy types to Python types
                if pd.isna(value):
                    row_dict[col] = None
                elif isinstance(value, (np.integer, np.floating)):
                    row_dict[col] = float(value) if isinstance(value, np.floating) else int(value)
                else:
                    row_dict[col] = str(value)
            rows.append(row_dict)

        # Create output directory
        os.makedirs(CONVERTED_DIR, exist_ok=True)

        # Convert to requested formats
        output_files = {}
        df = pd.DataFrame(rows)
        
        for format in output_formats:
            output_path = os.path.join(CONVERTED_DIR, f"{file_id}.{format}")
            if format == "xlsx":
                df.to_excel(output_path, index=False)
            elif format == "csv":
                df.to_csv(output_path, index=False)
            elif format == "xml":
                create_tally_xml(df, output_path)
            output_files[format] = output_path

        return {
            "status": "success",
            "message": "File converted successfully",
            "file_id": file_id,
            "headers": headers,
            "rows": rows,
            "converted_files": {
                format: f"/download/{os.path.basename(path)}"
                for format, path in output_files.items()
            }
        }

    except Exception as e:
        logger.error(f"Error converting file: {str(e)}")
        logger.error(traceback.format_exc())
        raise ConversionError(str(e))

@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    password: str = Form(None)
):
    """
    Upload a file and process it
    """
    try:
        # Create uploads directory if it doesn't exist
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        
        # Generate unique file ID
        file_id = str(uuid.uuid4())
        
        # Get file extension
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        # Create temporary file path
        temp_path = os.path.join(UPLOAD_DIR, f"temp_{file_id}{file_extension}")
        
        # Create final file path
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_extension}")
        
        # Save uploaded file to temporary location
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Validate file is a PDF if it has .pdf extension
        if file_extension == '.pdf':
            try:
                # Try PyMuPDF first
                doc = None
                try:
                    doc = fitz.open(temp_path)
                    if doc.needs_pass:
                        if not password:
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                            return JSONResponse(
                                status_code=401,
                                content={
                                    "requires_password": True,
                                    "message": "PDF is password protected. Please provide a password."
                                }
                            )
                        if not doc.authenticate(password):
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                            return JSONResponse(
                                status_code=401,
                                content={
                                    "requires_password": True,
                                    "message": "Incorrect password. Please try again."
                                }
                            )
                    # Validate PDF has pages
                    if doc.page_count == 0:
                        raise ValueError("PDF file has no pages")
                finally:
                    if doc:
                        doc.close()
                
            except Exception as e:
                logger.warning(f"PyMuPDF validation failed: {str(e)}, trying PyPDF2")
                # Try PyPDF2 as fallback
                try:
                    with open(temp_path, 'rb') as pdf_file:
                        pdf_reader = PdfReader(pdf_file)
                        if pdf_reader.is_encrypted:
                            if not password:
                                if os.path.exists(temp_path):
                                    os.remove(temp_path)
                                return JSONResponse(
                                    status_code=401,
                                    content={
                                        "requires_password": True,
                                        "message": "PDF is password protected. Please provide a password."
                                    }
                                )
                            try:
                                pdf_reader.decrypt(password)
                            except:
                                if os.path.exists(temp_path):
                                    os.remove(temp_path)
                                return JSONResponse(
                                    status_code=401,
                                    content={
                                        "requires_password": True,
                                        "message": "Incorrect password. Please try again."
                                    }
                                )
                        # Validate PDF has pages
                        if len(pdf_reader.pages) == 0:
                            raise ValueError("PDF file has no pages")
                except Exception as e:
                    logger.error(f"PDF validation failed: {str(e)}")
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    raise HTTPException(
                        status_code=400,
                        detail="Invalid or corrupted PDF file"
                    )
        
        # If we get here, either the file is not a PDF or it's a valid PDF
        # Move the temp file to final location
        os.rename(temp_path, file_path)
        
        # Return success response
        return {
            "file_id": file_id,
            "message": "File uploaded successfully"
        }
        
    except Exception as e:
        # Clean up temporary file if it exists
        if os.path.exists(temp_path):
            os.remove(temp_path)
        logger.error(f"Upload error: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

@app.post("/convert/{file_id}")
async def convert_uploaded_file(file_id: str):
    """Convert a file and return the extracted data"""
    try:
        logger.info(f"Starting conversion for file_id: {file_id}")
        
        # Find the original file
        original_file = None
        upload_dir_contents = os.listdir(UPLOAD_DIR)
        logger.info(f"Files in upload directory: {upload_dir_contents}")
        
        for filename in upload_dir_contents:
            if filename.startswith(file_id):
                original_file = os.path.join(UPLOAD_DIR, filename)
                logger.info(f"Found original file: {original_file}")
                break
        
        if not original_file:
            logger.error(f"Original file not found for file_id: {file_id}")
            raise HTTPException(status_code=404, detail="Original file not found")

        # Extract data from the file
        try:
            if original_file.lower().endswith('.pdf'):
                logger.info("Processing PDF file")
                df = extract_tables_from_pdf(original_file)
            elif original_file.lower().endswith(('.png', '.jpg', '.jpeg')):
                logger.info("Processing image file")
                df = process_image_ocr(original_file)
            else:
                raise ValueError(f"Unsupported file format: {original_file}")

            if df.empty:
                raise ValueError("No data could be extracted from the file")

            # Convert DataFrame to dictionary format while preserving numeric types
            headers = df.columns.tolist()
            rows = []
            for _, row in df.iterrows():
                row_dict = {}
                for col in headers:
                    value = row[col]
                    # Convert numpy types to Python types
                    if pd.isna(value):
                        row_dict[col] = None
                    elif isinstance(value, (np.integer, np.floating)):
                        row_dict[col] = float(value) if isinstance(value, np.floating) else int(value)
                    elif isinstance(value, pd.Timestamp):
                        row_dict[col] = value.strftime('%Y-%m-%d')
                    else:
                        row_dict[col] = str(value)
                rows.append(row_dict)

            # Create output directory
            os.makedirs(CONVERTED_DIR, exist_ok=True)

            # Save as JSON for future use
            json_data = {
                "headers": headers,
                "rows": rows
            }
            json_path = os.path.join(CONVERTED_DIR, f"{file_id}.json")
            with open(json_path, 'w') as f:
                json.dump(json_data, f)
            
            # Also save as Excel for compatibility
            excel_path = os.path.join(CONVERTED_DIR, f"{file_id}.xlsx")
            df.to_excel(excel_path, index=False, engine='openpyxl')
            
            return JSONResponse({
                "status": "success",
                "message": "File converted successfully",
                "file_id": file_id,
                "headers": headers,
                "rows": rows,
                "files": {
                    "json": f"/download/{file_id}.json",
                    "excel": f"/download/{file_id}.xlsx"
                }
            })

        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            logger.error(traceback.format_exc())
            raise HTTPException(
                status_code=500,
                detail=f"Error processing file: {str(e)}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in convert_uploaded_file: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error converting file: {str(e)}"
        )

@app.get("/api/convert/{file_id}/{format}")
async def convert_to_format(file_id: str, format: str):
    """Convert a file to a specific format"""
    try:
        logger.info(f"Starting format conversion for file_id: {file_id} to {format}")
        
        # Check if we have already converted data
        converted_file = os.path.join(CONVERTED_DIR, f"{file_id}.json")
        if not os.path.exists(converted_file):
            # If not, we need to convert first
            data = await convert_uploaded_file(file_id)
        else:
            # Use existing converted data
            with open(converted_file, 'r') as f:
                data = json.load(f)
        
        # Create DataFrame from the data
        if 'data' in data and 'rows' in data['data']:
            df = pd.DataFrame(data['data']['rows'])
        elif 'rows' in data:
            df = pd.DataFrame(data['rows'])
        else:
            raise ValueError("Invalid data format")
        
        # Create output filename
        output_filename = f"{file_id}.{format}"
        output_path = os.path.join(CONVERTED_DIR, output_filename)
        
        # Ensure the converted directory exists
        os.makedirs(CONVERTED_DIR, exist_ok=True)
        
        try:
            # Convert to the requested format
            if format == "xlsx":
                df.to_excel(output_path, index=False, engine='openpyxl')
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif format == "csv":
                df.to_csv(output_path, index=False)
                media_type = "text/csv"
            elif format == "xml":
                create_tally_xml(df, output_path)
                media_type = "application/xml"
            else:
                raise HTTPException(status_code=400, detail="Unsupported format")

            logger.info(f"Successfully converted to {format}")
            
            if not os.path.exists(output_path):
                logger.error(f"Output file was not created: {output_path}")
                raise HTTPException(
                    status_code=500,
                    detail="Failed to create output file"
                )

            # Return the converted file with proper headers
            headers = {
                'Content-Disposition': f'attachment; filename="converted-file.{format}"',
                'Access-Control-Expose-Headers': 'Content-Disposition'
            }
            
            return FileResponse(
                path=output_path,
                media_type=media_type,
                filename=f"converted-file.{format}",
                headers=headers
            )

        except Exception as e:
            logger.error(f"Error converting to format: {str(e)}")
            logger.error(traceback.format_exc())
            raise HTTPException(
                status_code=500,
                detail=f"Error converting to format: {str(e)}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Unexpected error: {str(e)}"
        )

@app.post("/convert")
async def convert_new_file(
    file: UploadFile = File(...),
    output_formats: List[str] = ["xlsx", "csv", "xml"]
):
    """Convert a new file upload"""
    try:
        # Save uploaded file
        file_path, file_id = await save_upload_file(file)
        
        # Convert the file
        result = convert_file(file_path, file_id, output_formats)
        
        # Clean up uploaded file
        os.remove(file_path)
        
        return JSONResponse(result)
        
    except ConversionError as e:
        return JSONResponse(
            status_code=400,
            content={"status": "error", "message": str(e)}
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Internal server error: {str(e)}"}
        )

@app.get("/download/{filename}")
async def download_file(filename: str):
    """Download a file from the corrected directory"""
    try:
        file_path = Path("corrected") / filename
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
            
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=get_media_type(filename)
        )
        
    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

def get_media_type(filename: str) -> str:
    """Get the media type based on file extension"""
    ext = filename.split(".")[-1].lower()
    media_types = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
        "xml": "application/xml"
    }
    return media_types.get(ext, "application/octet-stream")

def ensure_directory(path: str):
    if not os.path.exists(path):
        os.makedirs(path)

def save_as_xlsx(data: TableData, filepath: str):
    """Create a professionally formatted Excel file for bank statements"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from datetime import datetime
    import locale
    
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Statement"
        
        # Convert data to DataFrame first for easier manipulation
        rows_data = []
        for row in data.rows:
            row_dict = {header: row.cells[header].value for header in data.headers}
            rows_data.append(row_dict)
        
        df = pd.DataFrame(rows_data, columns=data.headers)
        
        # Clean and standardize column names
        column_mapping = {}
        standard_columns = {
            'date': ['date', 'transaction_date', 'txn_date', 'value_date', 'tran_date'],
            'description': ['description', 'narration', 'particulars', 'details', 'remarks'],
            'debit': ['debit', 'withdrawal', 'dr', 'debit_amount', 'withdrawal_amount'],
            'credit': ['credit', 'deposit', 'cr', 'credit_amount', 'deposit_amount'],
            'balance': ['balance', 'running_balance', 'closing_balance', 'available_balance'],
            'reference': ['reference', 'ref_no', 'utr', 'cheque_no', 'transaction_id'],
            'amount': ['amount', 'transaction_amount']
        }
        
        # Map existing columns to standard names
        final_headers = []
        for header in data.headers:
            header_lower = str(header).lower().replace(' ', '_')
            mapped = False
            for std_name, variations in standard_columns.items():
                if header_lower in variations or any(var in header_lower for var in variations):
                    column_mapping[header] = std_name.title()
                    if std_name.title() not in final_headers:
                        final_headers.append(std_name.title())
                    mapped = True
                    break
            if not mapped:
                clean_header = str(header).replace('_', ' ').title()
                column_mapping[header] = clean_header
                final_headers.append(clean_header)
        
        # Apply column mapping
        df_renamed = df.rename(columns=column_mapping)
        
        # Ensure standard column order
        preferred_order = ['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Amount', 'Balance']
        ordered_headers = []
        for col in preferred_order:
            if col in df_renamed.columns:
                ordered_headers.append(col)
        
        # Add any remaining columns
        for col in df_renamed.columns:
            if col not in ordered_headers:
                ordered_headers.append(col)
        
        df_final = df_renamed[ordered_headers]
        
        # Clean and format data
        for col in df_final.columns:
            if col == 'Date':
                # Standardize date format
                df_final[col] = pd.to_datetime(df_final[col], errors='coerce', infer_datetime_format=True)
                df_final[col] = df_final[col].dt.strftime('%d-%m-%Y')
                df_final[col] = df_final[col].replace('NaT', '')
            elif col in ['Debit', 'Credit', 'Amount', 'Balance']:
                # Clean and format numeric columns
                df_final[col] = df_final[col].astype(str).str.replace(r'[₹,\s]', '', regex=True)
                df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0)
            else:
                # Clean text columns
                df_final[col] = df_final[col].astype(str).str.strip()
                df_final[col] = df_final[col].replace('nan', '')
        
        # Define styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Calibri', size=11)
        data_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        date_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add title row
        ws['A1'] = 'Bank Statement Analysis'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:' + get_column_letter(len(ordered_headers)) + '1')
        
        # Add metadata
        ws['A2'] = f'Generated on: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
        ws['A2'].font = Font(name='Calibri', size=10, italic=True)
        ws.merge_cells('A2:' + get_column_letter(len(ordered_headers)) + '2')
        
        if not df_final.empty:
            ws['A3'] = f'Period: {df_final["Date"].iloc[0]} to {df_final["Date"].iloc[-1]}'
            ws['A3'].font = Font(name='Calibri', size=10, italic=True)
            ws.merge_cells('A3:' + get_column_letter(len(ordered_headers)) + '3')
        
        # Add headers starting from row 5
        header_row = 5
        for col_idx, header in enumerate(ordered_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data
        for row_idx, (_, row) in enumerate(df_final.iterrows(), header_row + 1):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = border
                
                # Apply specific formatting based on column type
                if col_name == 'Date':
                    cell.alignment = date_alignment
                elif col_name in ['Debit', 'Credit', 'Amount', 'Balance']:
                    cell.alignment = number_alignment
                    if isinstance(value, (int, float)) and value != 0:
                        cell.number_format = '#,##0.00'
                    # Color coding for amounts
                    if col_name == 'Debit' and value > 0:
                        cell.font = Font(name='Calibri', size=11, color='D32F2F')  # Red for debits
                    elif col_name == 'Credit' and value > 0:
                        cell.font = Font(name='Calibri', size=11, color='388E3C')  # Green for credits
                else:
                    cell.alignment = data_alignment
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(ordered_headers, 1):
            column_letter = get_column_letter(col_idx)
            
            # Calculate optimal width
            max_length = len(str(header))
            for row_idx in range(header_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width with reasonable limits
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add data validation for date columns
        if 'Date' in ordered_headers:
            date_col = ordered_headers.index('Date') + 1
            date_validation = DataValidation(
                type="date",
                formula1=datetime(2000, 1, 1),
                formula2=datetime(2050, 12, 31),
                showErrorMessage=True,
                errorTitle="Invalid Date",
                error="Please enter a valid date"
            )
            date_range = f"{get_column_letter(date_col)}{header_row + 1}:{get_column_letter(date_col)}{ws.max_row}"
            date_validation.add(date_range)
            ws.add_data_validation(date_validation)
        
        # Add summary section if we have balance data
        if 'Balance' in ordered_headers and not df_final.empty:
            summary_row = ws.max_row + 3
            
            # Summary title
            ws.cell(row=summary_row, column=1, value="Summary").font = Font(name='Calibri', size=12, bold=True)
            
            # Calculate summary statistics
            if 'Debit' in df_final.columns:
                total_debits = df_final['Debit'].sum()
                ws.cell(row=summary_row + 1, column=1, value="Total Debits:")
                ws.cell(row=summary_row + 1, column=2, value=total_debits).number_format = '#,##0.00'
            
            if 'Credit' in df_final.columns:
                total_credits = df_final['Credit'].sum()
                ws.cell(row=summary_row + 2, column=1, value="Total Credits:")
                ws.cell(row=summary_row + 2, column=2, value=total_credits).number_format = '#,##0.00'
            
            # Opening and closing balance
            if 'Balance' in df_final.columns:
                opening_balance = df_final['Balance'].iloc[0] if len(df_final) > 0 else 0
                closing_balance = df_final['Balance'].iloc[-1] if len(df_final) > 0 else 0
                
                ws.cell(row=summary_row + 3, column=1, value="Opening Balance:")
                ws.cell(row=summary_row + 3, column=2, value=opening_balance).number_format = '#,##0.00'
                
                ws.cell(row=summary_row + 4, column=1, value="Closing Balance:")
                ws.cell(row=summary_row + 4, column=2, value=closing_balance).number_format = '#,##0.00'
        
        # Freeze panes (freeze header row)
        ws.freeze_panes = f'A{header_row + 1}'
        
        # Add auto-filter
        if ws.max_row > header_row:
            ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(ordered_headers))}{ws.max_row}'
        
        # Save the workbook
        wb.save(filepath)
        
    except Exception as e:
        # Fallback to basic DataFrame export if formatting fails
        logger.warning(f"Advanced Excel formatting failed: {str(e)}, using basic export")
        rows_data = []
        for row in data.rows:
            row_dict = {header: row.cells[header].value for header in data.headers}
            rows_data.append(row_dict)
        
        df = pd.DataFrame(rows_data, columns=data.headers)
        df.to_excel(filepath, index=False, engine='openpyxl')

def save_as_csv(data: TableData, filepath: str):
    """Create a clean CSV file for bank statements"""
    try:
        # Convert to pandas DataFrame
        rows_data = []
        for row in data.rows:
            row_dict = {header: row.cells[header].value for header in data.headers}
            rows_data.append(row_dict)
        
        df = pd.DataFrame(rows_data, columns=data.headers)
        
        # Clean and standardize column names similar to Excel
        column_mapping = {}
        standard_columns = {
            'date': ['date', 'transaction_date', 'txn_date', 'value_date', 'tran_date'],
            'description': ['description', 'narration', 'particulars', 'details', 'remarks'],
            'debit': ['debit', 'withdrawal', 'dr', 'debit_amount', 'withdrawal_amount'],
            'credit': ['credit', 'deposit', 'cr', 'credit_amount', 'deposit_amount'],
            'balance': ['balance', 'running_balance', 'closing_balance', 'available_balance'],
            'reference': ['reference', 'ref_no', 'utr', 'cheque_no', 'transaction_id'],
            'amount': ['amount', 'transaction_amount']
        }
        
        # Map existing columns to standard names
        for header in data.headers:
            header_lower = str(header).lower().replace(' ', '_')
            mapped = False
            for std_name, variations in standard_columns.items():
                if header_lower in variations or any(var in header_lower for var in variations):
                    column_mapping[header] = std_name.title()
                    mapped = True
                    break
            if not mapped:
                clean_header = str(header).replace('_', ' ').title()
                column_mapping[header] = clean_header
        
        # Apply column mapping
        df_renamed = df.rename(columns=column_mapping)
        
        # Ensure standard column order
        preferred_order = ['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Amount', 'Balance']
        ordered_headers = []
        for col in preferred_order:
            if col in df_renamed.columns:
                ordered_headers.append(col)
        
        # Add any remaining columns
        for col in df_renamed.columns:
            if col not in ordered_headers:
                ordered_headers.append(col)
        
        df_final = df_renamed[ordered_headers]
        
        # Clean and format data
        for col in df_final.columns:
            if col == 'Date':
                # Standardize date format
                df_final[col] = pd.to_datetime(df_final[col], errors='coerce', infer_datetime_format=True)
                df_final[col] = df_final[col].dt.strftime('%d-%m-%Y')
                df_final[col] = df_final[col].replace('NaT', '')
            elif col in ['Debit', 'Credit', 'Amount', 'Balance']:
                # Clean and format numeric columns
                df_final[col] = df_final[col].astype(str).str.replace(r'[₹,\s]', '', regex=True)
                df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0)
            else:
                # Clean text columns
                df_final[col] = df_final[col].astype(str).str.strip()
                df_final[col] = df_final[col].replace('nan', '')
        
        # Save to CSV with proper encoding
        df_final.to_csv(filepath, index=False, encoding='utf-8')
        
    except Exception as e:
        # Fallback to basic DataFrame export if formatting fails
        logger.warning(f"Advanced CSV formatting failed: {str(e)}, using basic export")
        rows_data = []
        for row in data.rows:
            row_dict = {header: row.cells[header].value for header in data.headers}
            rows_data.append(row_dict)
        
        df = pd.DataFrame(rows_data, columns=data.headers)
        df.to_csv(filepath, index=False)

def save_as_xml(data: TableData, filepath: str):
    root = ET.Element("ENVELOPE")
    header = ET.SubElement(root, "HEADER")
    ET.SubElement(header, "VERSION").text = "1"
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
    
    body = ET.SubElement(root, "BODY")
    importdata = ET.SubElement(body, "IMPORTDATA")
    requestdesc = ET.SubElement(importdata, "REQUESTDESC")
    reportname = ET.SubElement(requestdesc, "REPORTNAME").text = "Custom"
    
    requestdata = ET.SubElement(importdata, "REQUESTDATA")
    
    # Convert rows to XML structure
    for row in data.rows:
        tallymessage = ET.SubElement(requestdata, "TALLYMESSAGE")
        for header in data.headers:
            ET.SubElement(tallymessage, str(header)).text = str(row.cells[header].value)
    
    # Save XML file
    tree = ET.ElementTree(root)
    tree.write(filepath, encoding='utf-8', xml_declaration=True)

def log_changes(file_id: str, edit_history: List[EditHistory]):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_dir = "logs"
    ensure_directory(log_dir)
    
    log_file = os.path.join(log_dir, f"{file_id}_{timestamp}_changes.json")
    with open(log_file, "w") as f:
        json.dump([edit.dict() for edit in edit_history], f, indent=2)

@app.post("/api/save-edits")
async def save_edits(payload: SavePayload):
    try:
        # Ensure the corrected directory exists
        corrected_dir = "corrected"
        ensure_directory(corrected_dir)
        
        # Save in different formats
        base_path = os.path.join(corrected_dir, f"{payload.fileId}_corrected")
        save_as_xlsx(payload.modifiedData, f"{base_path}.xlsx")
        save_as_csv(payload.modifiedData, f"{base_path}.csv")
        save_as_xml(payload.modifiedData, f"{base_path}.xml")
        
        # Log the changes
        log_changes(payload.fileId, payload.editHistory)
        
        return {
            "success": True,
            "downloads": {
                "xlsx": f"/api/download/{payload.fileId}/xlsx",
                "csv": f"/api/download/{payload.fileId}/csv",
                "xml": f"/api/download/{payload.fileId}/xml"
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/download/{file_id}/{format}")
async def download_converted_file(file_id: str, format: str):
    try:
        # Check if the file exists in converted directory
        converted_dir = "converted"
        filename = f"{file_id}.{format}"
        file_path = os.path.join(converted_dir, filename)
        
        if not os.path.exists(file_path):
            # Try to generate the file from the JSON data
            json_path = os.path.join(converted_dir, f"{file_id}.json")
            if not os.path.exists(json_path):
                raise HTTPException(status_code=404, detail="File not found")
            
            with open(json_path, 'r') as f:
                data = json.load(f)
            
            df = pd.DataFrame(data)
            
            if format == 'xlsx':
                output_path = os.path.join(converted_dir, f"{file_id}.xlsx")
                df.to_excel(output_path, index=False, engine='openpyxl')
                file_path = output_path
            elif format == 'csv':
                output_path = os.path.join(converted_dir, f"{file_id}.csv")
                df.to_csv(output_path, index=False)
                file_path = output_path
            elif format == 'xml':
                output_path = os.path.join(converted_dir, f"{file_id}.xml")
                # Convert DataFrame to XML
                xml_data = df.to_xml(index=False)
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(xml_data)
                file_path = output_path
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")
        
        return FileResponse(
            file_path,
            filename=f"converted_file.{format}",
            media_type={
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'csv': 'text/csv',
                'xml': 'application/xml'
            }.get(format, 'application/octet-stream')
        )
    except Exception as e:
        logging.error(f"Error downloading file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/files")
async def list_files():
    """List all uploaded files"""
    try:
        files = []
        for filename in os.listdir(UPLOAD_DIR):
            file_path = os.path.join(UPLOAD_DIR, filename)
            if os.path.isfile(file_path):
                file_id = os.path.splitext(filename)[0]
                file_stats = os.stat(file_path)
                mime_type, _ = mimetypes.guess_type(filename)
                
                files.append({
                    "id": file_id,
                    "name": filename,
                    "type": mime_type or "application/octet-stream",
                    "size": file_stats.st_size,
                    "uploadedAt": datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                    "url": f"/files/{filename}"
                })
        
        return {
            "status": "success",
            "files": sorted(files, key=lambda x: x["uploadedAt"], reverse=True)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/files/{filename}")
async def serve_file(filename: str):
    """Serve an uploaded file"""
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    mime_type, _ = mimetypes.guess_type(filename)
    return FileResponse(
        file_path,
        media_type=mime_type or "application/octet-stream",
        filename=filename
    )

def validate_table(data: List[List[str]]) -> List[Dict]:
    """Validate table data and return validation results"""
    validation_results = []
    
    for row_idx, row in enumerate(data):
        for col_idx, cell in enumerate(row):
            # Skip header row
            if row_idx == 0:
                continue
                
            # Validate empty mandatory fields
            if not cell.strip() and col_idx in [0, 1, 2]:  # Assuming first 3 columns are mandatory
                validation_results.append({
                    "row": row_idx,
                    "column": col_idx,
                    "type": "error",
                    "severity": "critical",
                    "message": "Mandatory field cannot be empty"
                })
            
            # Validate numeric values
            if col_idx in [3, 4]:  # Assuming columns 4 and 5 should be numeric
                try:
                    float(cell.replace(',', '').strip())
                except ValueError:
                    validation_results.append({
                        "row": row_idx,
                        "column": col_idx,
                        "type": "error",
                        "severity": "critical",
                        "message": "Value must be numeric"
                    })
            
            # Validate date format
            if col_idx == 2:  # Assuming column 3 is date
                try:
                    datetime.strptime(cell.strip(), '%Y-%m-%d')
                except ValueError:
                    validation_results.append({
                        "row": row_idx,
                        "column": col_idx,
                        "type": "error",
                        "severity": "warning",
                        "message": "Invalid date format (should be YYYY-MM-DD)"
                    })
            
            # Check for common OCR errors
            if any(char in cell for char in ['O0', 'l1', 'S5']):
                validation_results.append({
                    "row": row_idx,
                    "column": col_idx,
                    "type": "warning",
                    "severity": "info",
                    "message": "Possible OCR confusion (O/0, l/1, S/5)"
                })
    
    return validation_results

@app.post("/validate/{file_id}")
async def validate_data(file_id: str, request: Request):
    """Validate converted data for a specific file"""
    try:
        # Get the request body
        data = await request.json()
        
        if not data:
            raise HTTPException(status_code=400, detail="No data provided for validation")

        # Get the file path
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}.pdf")
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Extract data and formats from request
        table_data = data.get('data', [])
        formats = data.get('formats', [])
        password = data.get('password')

        if not table_data:
            raise HTTPException(status_code=422, detail="No table data provided")
        
        # Ensure table_data is a list
        if not isinstance(table_data, list):
            table_data = [table_data]
        
        # Convert to DataFrame for validation
        try:
            df = pd.DataFrame(table_data)
        except Exception as e:
            logger.error(f"Error converting data to DataFrame: {str(e)}")
            raise HTTPException(
                status_code=422,
                detail=f"Invalid data format: {str(e)}"
            )
        
        # Clean up data types
        try:
            # Convert date columns if they exist
            date_columns = [col for col in df.columns if any(term in col.lower() for term in ['date', 'dt', 'time'])]
            for col in date_columns:
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                except Exception as e:
                    logger.warning(f"Could not convert column {col} to date: {str(e)}")
            
            # Convert numeric columns if they exist
            numeric_columns = [col for col in df.columns if any(term in col.lower() for term in ['amount', 'balance', 'total', 'price', 'quantity'])]
            for col in numeric_columns:
                try:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace('₹', '').str.replace(',', ''), errors='coerce')
                except Exception as e:
                    logger.warning(f"Could not convert column {col} to numeric: {str(e)}")
            
            # Fill NaN values
            df = df.fillna('')
            
        except Exception as e:
            logger.error(f"Error cleaning data: {str(e)}")
            raise HTTPException(
                status_code=422,
                detail=f"Error cleaning data: {str(e)}"
            )
        
        try:
            # Convert DataFrame back to records for validation
            records = df.to_dict('records')
            
            # Perform validation checks
            validation_results = validate_table_data(records)
            financial_validation = validate_financial_data(records)
            
            # Get validation summary
            summary = get_validation_summary(records)
            
            # Save validated data
            try:
                os.makedirs("corrected", exist_ok=True)
                validated_path = os.path.join("corrected", f"{file_id}_validated.json")
                with open(validated_path, 'w') as f:
                    json.dump({
                        'data': records,
                        'formats': formats,
                        'validation': {
                            'results': validation_results,
                            'financial': financial_validation,
                            'summary': summary
                        }
                    }, f, indent=2)
            except Exception as e:
                logger.error(f"Error saving validated data: {str(e)}")
                # Don't fail the validation if saving fails
                pass
            
            return JSONResponse(
                status_code=200,
                content={
                    "status": "success",
                    "message": "Validation completed successfully",
                    "validationResults": {
                        "results": validation_results,
                        "financial": financial_validation,
                        "summary": summary
                    }
                }
            )
            
        except Exception as e:
            logger.error(f"Error during validation: {str(e)}")
            logger.error(traceback.format_exc())
            raise HTTPException(
                status_code=422,
                detail=f"Validation error: {str(e)}"
            )
            
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Unexpected error during validation: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Unexpected error: {str(e)}"
        )

@app.post("/export/{file_id}/{format}")
async def export_data(
    file_id: str,
    format: str,
    request: Request,
    background_tasks: BackgroundTasks
):
    """Enhanced export with multi-language support"""
    try:
        # Parse request body for language and localization settings
        body = await request.json()
        language = body.get('language', 'en')
        localization = body.get('localization', {})
        data = body.get('data', [])
        
        logger.info(f"Exporting {format} for file {file_id} in language {language}")
        
        # Define language-specific settings
        language_settings = {
            'en': {
                'date_format': '%m/%d/%Y',
                'number_format': 'US',
                'currency_symbol': '$',
                'headers': {
                    'date': 'Date',
                    'description': 'Description',
                    'debit': 'Debit',
                    'credit': 'Credit',
                    'balance': 'Balance',
                    'reference': 'Reference',
                    'amount': 'Amount'
                }
            },
            'hi': {
                'date_format': '%d/%m/%Y',
                'number_format': 'IN',
                'currency_symbol': '₹',
                'headers': {
                    'date': 'दिनांक',
                    'description': 'विवरण',
                    'debit': 'डेबिट',
                    'credit': 'क्रेडिट',
                    'balance': 'शेष',
                    'reference': 'संदर्भ',
                    'amount': 'राशि'
                }
            },
            'mr': {
                'date_format': '%d/%m/%Y',
                'number_format': 'IN',
                'currency_symbol': '₹',
                'headers': {
                    'date': 'दिनांक',
                    'description': 'तपशील',
                    'debit': 'डेबिट',
                    'credit': 'क्रेडिट',
                    'balance': 'शिल्लक',
                    'reference': 'संदर्भ',
                    'amount': 'रक्कम'
                }
            }
        }
        
        current_lang_settings = language_settings.get(language, language_settings['en'])
        
        # Override with custom localization if provided
        if localization:
            if 'dateFormat' in localization:
                current_lang_settings['date_format'] = localization['dateFormat'].replace('MM', '%m').replace('DD', '%d').replace('YYYY', '%Y')
            if 'currency' in localization:
                current_lang_settings['currency_symbol'] = localization['currency']
        
        # Determine data source paths
        export_dir = "exports"
        converted_dir = "converted"  
        corrected_dir = "corrected"
        
        ensure_directory(export_dir)
        
        # Try to find data in order of preference: corrected -> converted -> provided
        df = None
        data_source = "provided"
        
        # Check for validated/corrected data first
        validated_json_path = os.path.join(corrected_dir, f"{file_id}_validated.json")
        if os.path.exists(validated_json_path):
            try:
                with open(validated_json_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                df = pd.DataFrame(json_data)
                data_source = "validated"
                logger.info(f"Using validated data from {validated_json_path}")
            except Exception as e:
                logger.warning(f"Error reading validated data: {str(e)}")
        
        # Fallback to converted data
        if df is None:
            converted_json_path = os.path.join(converted_dir, f"{file_id}.json")
            if os.path.exists(converted_json_path):
                try:
                    with open(converted_json_path, 'r', encoding='utf-8') as f:
                        json_data = json.load(f)
                    df = pd.DataFrame(json_data)
                    data_source = "converted"
                    logger.info(f"Using converted data from {converted_json_path}")
                except Exception as e:
                    logger.warning(f"Error reading converted data: {str(e)}")
        
        # Final fallback to provided data
        if df is None and data:
            df = pd.DataFrame(data)
            data_source = "provided"
            logger.info("Using provided data from request")
        
        if df is None or df.empty:
            raise HTTPException(status_code=404, detail="No data found for export")
        
        # Apply language-specific formatting to DataFrame
        def localize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
            df_localized = df.copy()
            
            # Rename columns to localized headers
            column_mapping = {}
            for col in df_localized.columns:
                col_lower = col.lower()
                for eng_key, local_header in current_lang_settings['headers'].items():
                    if eng_key in col_lower or col_lower == eng_key:
                        column_mapping[col] = local_header
                        break
            
            if column_mapping:
                df_localized = df_localized.rename(columns=column_mapping)
            
            # Format date columns
            for col in df_localized.columns:
                if 'date' in col.lower() or col in current_lang_settings['headers'].values():
                    try:
                        df_localized[col] = pd.to_datetime(df_localized[col], errors='coerce')
                        df_localized[col] = df_localized[col].dt.strftime(current_lang_settings['date_format'])
                    except:
                        pass
            
            # Format numeric columns with currency
            for col in df_localized.columns:
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in ['amount', 'debit', 'credit', 'balance']) or col in current_lang_settings['headers'].values():
                    try:
                        # Convert to numeric first
                        df_localized[col] = pd.to_numeric(df_localized[col], errors='coerce')
                        # Format with currency symbol for display
                        if current_lang_settings['number_format'] == 'IN':
                            df_localized[col] = df_localized[col].apply(
                                lambda x: f"{current_lang_settings['currency_symbol']}{x:,.2f}" if pd.notna(x) and x != 0 else ""
                            )
                        else:
                            df_localized[col] = df_localized[col].apply(
                                lambda x: f"{current_lang_settings['currency_symbol']}{x:,.2f}" if pd.notna(x) and x != 0 else ""
                            )
                    except:
                        pass
            
            return df_localized
        
        # Apply localization
        df_localized = localize_dataframe(df)
        
        # Generate export filename with language suffix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"{file_id}_{language}_{timestamp}"
        
        if format.lower() == 'xlsx':
            filename = f"{base_filename}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Use enhanced Excel formatting with localization
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Bank Statement ({language.upper()})"
            
            # Add title with language indicator
            title_text = {
                'en': 'Bank Statement Analysis',
                'hi': 'बैंक स्टेटमेंट विश्लेषण', 
                'mr': 'बँक स्टेटमेंट विश्लेषण'
            }
            ws['A1'] = title_text.get(language, title_text['en'])
            ws['A1'].font = Font(name='Calibri', size=16, bold=True)
            ws.merge_cells(f'A1:{get_column_letter(len(df_localized.columns))}1')
            
            # Add generation info
            ws['A2'] = f'Generated: {datetime.now().strftime(current_lang_settings["date_format"])} | Language: {language.upper()}'
            ws['A2'].font = Font(name='Calibri', size=10, italic=True)
            ws.merge_cells(f'A2:{get_column_letter(len(df_localized.columns))}2')
            
            # Add headers and data starting from row 4
            for col_idx, header in enumerate(df_localized.columns, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data
            for row_idx, (_, row) in enumerate(df_localized.iterrows(), 5):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col_idx in range(1, len(df_localized.columns) + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15
            
            wb.save(filepath)
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        elif format.lower() == 'csv':
            filename = f"{base_filename}.csv"
            filepath = os.path.join(export_dir, filename)
            df_localized.to_csv(filepath, index=False, encoding='utf-8-sig')  # Use utf-8-sig for Excel compatibility
            media_type = 'text/csv'
            
        elif format.lower() == 'json':
            filename = f"{base_filename}.json"
            filepath = os.path.join(export_dir, filename)
            
            # Create structured JSON with metadata
            export_data = {
                'metadata': {
                    'file_id': file_id,
                    'language': language,
                    'export_date': datetime.now().isoformat(),
                    'data_source': data_source,
                    'localization': current_lang_settings
                },
                'data': df_localized.to_dict('records')
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            media_type = 'application/json'
            
        elif format.lower() in ['xml', 'tally']:
            filename = f"{base_filename}.xml"
            filepath = os.path.join(export_dir, filename)
            
            if format.lower() == 'tally':
                # Use existing Tally XML function with localized headers
                create_tally_xml(df_localized, filepath)
            else:
                # Generic XML with localization
                root = ET.Element("BankStatementExport")
                root.set("language", language)
                root.set("exportDate", datetime.now().isoformat())
                
                metadata = ET.SubElement(root, "Metadata")
                ET.SubElement(metadata, "FileId").text = file_id
                ET.SubElement(metadata, "Language").text = language
                ET.SubElement(metadata, "DataSource").text = data_source
                
                data_elem = ET.SubElement(root, "Data")
                for _, row in df_localized.iterrows():
                    record = ET.SubElement(data_elem, "Record")
                    for col, value in row.items():
                        elem = ET.SubElement(record, "Field")
                        elem.set("name", str(col))
                        elem.text = str(value) if value is not None else ""
                
                tree = ET.ElementTree(root)
                tree.write(filepath, encoding='utf-8', xml_declaration=True)
            
            media_type = 'application/xml'
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")
        
        # Add cleanup task
        background_tasks.add_task(cleanup_old_exports, export_dir)
        
        # Return file with proper headers
        return FileResponse(
            filepath,
            filename=filename,
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Cache-Control": "no-cache",
                "X-Language": language,
                "X-Data-Source": data_source
            }
        )
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

def cleanup_old_exports(export_dir: str, max_age_days: int = 7):
    """Clean up export files older than max_age_days."""
    try:
        current_time = datetime.now()
        for file in os.listdir(export_dir):
            if file == "audit_log.json":
                continue
                
            file_path = os.path.join(export_dir, file)
            file_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
            
            if (current_time - file_modified).days > max_age_days:
                os.remove(file_path)
                logging.info(f"Cleaned up old export: {file}")
    except Exception as e:
        logging.error(f"Error cleaning up exports: {str(e)}")

@app.post("/bank-statement/upload")
async def upload_bank_statement(file: UploadFile = File(...)):
    """Upload and process a bank statement"""
    try:
        file_path, file_id = await save_upload_file(file)
        
        # Process the bank statement
        transactions = bank_matcher.load_bank_statement(file_path)
        
        return {
            "success": True,
            "message": "Bank statement processed successfully",
            "transaction_count": len(transactions)
        }
    except Exception as e:
        logger.error(f"Error processing bank statement: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error processing bank statement: {str(e)}"
        )

@app.post("/invoice-data/upload")
async def upload_invoice_data(file: UploadFile = File(...)):
    """Upload and process invoice data"""
    try:
        file_path, file_id = await save_upload_file(file)
        
        # Process the invoice data based on file type
        if file_path.lower().endswith('.pdf'):
            df = process_bank_statement(file_path)  # Reuse bank statement parser
        else:
            df = pd.read_csv(file_path)
        
        # Convert DataFrame to list of dictionaries
        invoice_data = df.to_dict('records')
        
        # Load into bank matcher
        transactions = bank_matcher.load_invoice_data(invoice_data)
        
        return {
            "success": True,
            "message": "Invoice data processed successfully",
            "transaction_count": len(transactions)
        }
    except Exception as e:
        logger.error(f"Error processing invoice data: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error processing invoice data: {str(e)}"
        )

@app.post("/bank-matcher/match")
async def match_transactions(
    date_tolerance_days: int = Query(5, description="Number of days to consider for matching dates"),
    amount_tolerance: float = Query(0.01, description="Amount difference tolerance for matching")
):
    """Find matches between bank transactions and invoices"""
    try:
        matches = bank_matcher.find_matches(
            date_tolerance_days=date_tolerance_days,
            amount_tolerance=amount_tolerance
        )
        
        unmatched = bank_matcher.get_unmatched_transactions()
        
        return {
            "success": True,
            "matches": matches,
            "unmatched": unmatched,
            "match_count": len(matches)
        }
    except Exception as e:
        logger.error(f"Error matching transactions: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error matching transactions: {str(e)}"
        )

@app.post("/bank-matcher/update-status")
async def update_match_status(
    match_id: str = Body(..., embed=True),
    status: str = Body(..., embed=True)
):
    """Update the status of a match"""
    try:
        success = bank_matcher.update_match_status(match_id, status)
        if not success:
            raise HTTPException(
                status_code=404,
                detail=f"Match with ID {match_id} not found"
            )
        
        return {
            "success": True,
            "message": f"Match status updated to {status}"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating match status: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error updating match status: {str(e)}"
        )

@app.post("/gst/add-invoice")
async def add_gst_invoice(invoice_data: Dict[str, Any]):
    """Add an invoice for GST processing"""
    try:
        invoice = gst_helper.add_invoice(invoice_data)
        return {
            "success": True,
            "message": "Invoice added successfully",
            "invoice": {
                "invoice_no": invoice.invoice_no,
                "total_amount": invoice.total_amount,
                "gst_details": {
                    "taxable_value": invoice.taxable_value,
                    "igst": invoice.igst,
                    "cgst": invoice.cgst,
                    "sgst": invoice.sgst
                }
            }
        }
    except Exception as e:
        logger.error(f"Error adding GST invoice: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error adding GST invoice: {str(e)}"
        )

@app.post("/gst/generate-gstr1")
async def generate_gstr1(period: str = Body(..., embed=True)):
    """Generate GSTR-1 format JSON"""
    try:
        # Check if there are any invoices
        if gst_helper.get_invoice_count() == 0:
            return {
                "success": True,
                "message": "No invoices found for GSTR-1 generation",
                "data": {
                    "gstin": "",
                    "fp": period,
                    "version": "GST3.0.4",
                    "hash": "hash",
                    "b2b": [],
                    "exp": []
                }
            }
        
        gstr1_data = gst_helper.generate_gstr1_json(period)
        
        # Ensure export directory exists
        os.makedirs(EXPORT_DIR, exist_ok=True)
        
        # Save to file
        file_path = os.path.join(EXPORT_DIR, f"GSTR1_{period}.json")
        with open(file_path, 'w') as f:
            json.dump(gstr1_data, f, indent=2)
        
        return FileResponse(
            file_path,
            media_type='application/json',
            filename=f"GSTR1_{period}.json"
        )
    except Exception as e:
        logger.error(f"Error generating GSTR-1: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error generating GSTR-1: {str(e)}"
        )

@app.post("/gst/generate-gstr3b")
async def generate_gstr3b(period: str = Body(..., embed=True)):
    """Generate GSTR-3B format JSON"""
    try:
        # Check if there are any invoices
        if gst_helper.get_invoice_count() == 0:
            return {
                "success": True,
                "message": "No invoices found for GSTR-3B generation",
                "data": {
                    "gstin": "",
                    "ret_period": period,
                    "sup_details": {
                        "osup_det": {"txval": 0, "iamt": 0, "camt": 0, "samt": 0, "csamt": 0},
                        "osup_zero": {"txval": 0, "iamt": 0, "csamt": 0},
                        "osup_nil_exmp": {"txval": 0},
                        "isup_rev": {"txval": 0, "iamt": 0, "camt": 0, "samt": 0, "csamt": 0}
                    }
                }
            }
        
        gstr3b_data = gst_helper.generate_gstr3b_json(period)
        
        # Ensure export directory exists
        os.makedirs(EXPORT_DIR, exist_ok=True)
        
        # Save to file
        file_path = os.path.join(EXPORT_DIR, f"GSTR3B_{period}.json")
        with open(file_path, 'w') as f:
            json.dump(gstr3b_data, f, indent=2)
        
        return FileResponse(
            file_path,
            media_type='application/json',
            filename=f"GSTR3B_{period}.json"
        )
    except Exception as e:
        logger.error(f"Error generating GSTR-3B: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error generating GSTR-3B: {str(e)}"
        )

@app.post("/gst/generate-excel")
async def generate_gst_excel(period: str = Body(..., embed=True)):
    """Generate GST Excel report"""
    try:
        # Ensure export directory exists
        os.makedirs(EXPORT_DIR, exist_ok=True)
        
        file_path = os.path.join(EXPORT_DIR, f"GST_Report_{period}.xlsx")
        
        # Generate the Excel report
        gst_helper.generate_excel_report(file_path)
        
        # Check if file was created successfully
        if not os.path.exists(file_path):
            raise HTTPException(
                status_code=500,
                detail="Failed to generate Excel report file"
            )
        
        return FileResponse(
            file_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"GST_Report_{period}.xlsx"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating GST Excel report: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error generating GST Excel report: {str(e)}"
        )

@app.post("/gst/validate-gstin")
async def validate_gstin(gstin: str = Body(..., embed=True)):
    """Validate GSTIN format"""
    try:
        is_valid = gst_helper.validate_gstin(gstin)
        return {
            "success": True,
            "is_valid": is_valid
        }
    except Exception as e:
        logger.error(f"Error validating GSTIN: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error validating GSTIN: {str(e)}"
        )

@app.post("/gst/clear-invoices")
async def clear_gst_invoices():
    """Clear all GST invoices from the helper"""
    try:
        gst_helper.clear_invoices()
        return {
            "success": True,
            "message": "All GST invoices cleared successfully"
        }
    except Exception as e:
        logger.error(f"Error clearing GST invoices: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error clearing GST invoices: {str(e)}"
        )

@app.get("/gst/invoice-count")
async def get_gst_invoice_count():
    """Get the number of invoices in the GST helper"""
    try:
        count = gst_helper.get_invoice_count()
        return {
            "success": True,
            "invoice_count": count
        }
    except Exception as e:
        logger.error(f"Error getting GST invoice count: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error getting GST invoice count: {str(e)}"
        )

@app.get("/file/{file_id}")
async def get_file(file_id: str, convert: bool = Query(False)):
    """Get file by ID and optionally convert it to table format"""
    try:
        logger.info(f"Processing file request: file_id={file_id}, convert={convert}")
        
        # Check all possible file extensions
        file_found = False
        file_path = None
        file_ext = None
        
        for ext in ['.pdf', '.png', '.jpg', '.jpeg']:
            test_path = os.path.join(UPLOAD_DIR, f"{file_id}{ext}")
            logger.info(f"Checking file path: {test_path}")
            if os.path.exists(test_path):
                file_path = test_path
                file_ext = ext
                file_found = True
                logger.info(f"File found: {file_path}")
                break
        
        if not file_found:
            logger.error(f"File not found for ID: {file_id}")
            # List available files for debugging
            try:
                available_files = os.listdir(UPLOAD_DIR)
                logger.info(f"Available files in upload directory: {available_files}")
            except Exception as e:
                logger.error(f"Could not list upload directory: {str(e)}")
            raise HTTPException(status_code=404, detail="File not found")
        
        if not convert:
            # Return raw file if conversion not requested
            mime_type, _ = mimetypes.guess_type(file_path)
            if not mime_type:
                mime_type = 'application/octet-stream'
            
            return FileResponse(
                file_path,
                media_type=mime_type,
                filename=os.path.basename(file_path)
            )
        else:
            # Convert file to table format
            logger.info(f"Starting conversion for file: {file_path}")
            try:
                df = None
                if file_ext.lower() == '.pdf':
                    logger.info("Processing PDF file")
                    df = extract_tables_from_pdf(file_path)
                else:
                    logger.info("Processing image file")
                    df = process_image_ocr(file_path)
                
                logger.info(f"Conversion result: df is None: {df is None}, df is empty: {df.empty if df is not None else 'N/A'}")
                
                # Convert DataFrame to response format
                if df is not None and not df.empty:
                    logger.info(f"DataFrame shape: {df.shape}")
                    logger.info(f"DataFrame columns: {list(df.columns)}")
                    
                    try:
                        rows = df.to_dict('records')
                        headers = df.columns.tolist()
                        
                        logger.info(f"Successfully converted to {len(rows)} rows with {len(headers)} columns")
                        
                        return {
                            "success": True,
                            "data": {
                                "rows": rows,
                                "headers": headers
                            }
                        }
                    except Exception as e:
                        logger.error(f"Error converting DataFrame to dict: {str(e)}")
                        logger.error(traceback.format_exc())
                        raise HTTPException(
                            status_code=500,
                            detail=f"Error formatting conversion results: {str(e)}"
                        )
                else:
                    logger.warning("No data extracted from file or DataFrame is empty")
                    raise HTTPException(
                        status_code=422,
                        detail="No tabular data could be extracted from the file. The file may not contain tables or the format may not be supported."
                    )
                    
            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Error converting file {file_id}: {str(e)}")
                logger.error(traceback.format_exc())
                raise HTTPException(
                    status_code=500,
                    detail=f"Error converting file: {str(e)}"
                )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in get_file for {file_id}: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Unexpected server error: {str(e)}"
        )

@app.get("/audit-logs")
async def get_audit_logs(
    limit: Optional[int] = None,
    action_type: Optional[str] = None,
    user_id: Optional[str] = None
):
    """Get audit logs with optional filtering"""
    return audit_logger.get_logs(limit=limit, action_type=action_type, user_id=user_id)

@app.post("/audit-logs")
async def create_audit_log(
    action_type: str = Body(...),
    summary: str = Body(...),
    user_id: Optional[str] = Body(None),
    metadata: Optional[Dict[str, Any]] = Body(None)
):
    """Create a new audit log entry"""
    return audit_logger.log_action(
        action_type=action_type,
        summary=summary,
        user_id=user_id,
        metadata=metadata
    )

@app.get("/missed-rows/{file_id}")
async def get_missed_rows(file_id: str):
    """Get rows that failed to parse from bank statement"""
    try:
        # Get the file path
        file_path = None
        for ext in ['.pdf', '.png', '.jpg', '.jpeg']:
            path = os.path.join(UPLOAD_DIR, f"{file_id}{ext}")
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Process the file and get missed rows
        parser = BankStatementParser()
        _, missed_rows = parser.process_bank_statement(file_path)
        
        return {
            "success": True,
            "missed_rows": missed_rows
        }
        
    except Exception as e:
        logger.error(f"Error getting missed rows: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error getting missed rows: {str(e)}"
        )

@app.post("/auto-fix/{file_id}")
async def apply_auto_fixes(file_id: str):
    """Apply automatic fixes to validation issues"""
    try:
        # Get the validated data file
        validated_path = os.path.join("corrected", f"{file_id}_validated.json")
        if not os.path.exists(validated_path):
            raise HTTPException(status_code=404, detail="Validated data not found")
        
        with open(validated_path, 'r') as f:
            validated_data = json.load(f)
        
        # Apply automatic fixes based on validation results
        if 'validation' in validated_data and 'results' in validated_data['validation']:
            fixed_count = 0
            for result in validated_data['validation']['results']:
                for issue in result.get('issues', []):
                    if issue.get('suggestedValue') and issue.get('autoFixable', False):
                        # Apply the suggested fix
                        row_index = result['row']
                        field = issue['field']
                        if row_index < len(validated_data['data']):
                            validated_data['data'][row_index][field] = issue['suggestedValue']
                            fixed_count += 1
            
            # Save the fixed data
            with open(validated_path, 'w') as f:
                json.dump(validated_data, f, indent=2)
            
            return {
                "status": "success",
                "message": f"Applied {fixed_count} auto-fixes successfully",
                "fixed_count": fixed_count
            }
        else:
            return {
                "status": "success", 
                "message": "No auto-fixable issues found",
                "fixed_count": 0
            }
            
    except Exception as e:
        logger.error(f"Error applying auto-fixes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Auto-fix failed: {str(e)}")

@app.post("/update-cell/{file_id}")
async def update_cell(file_id: str, request: Request):
    """Update a specific cell in the validated data"""
    try:
        data = await request.json()
        row_index = data.get('rowIndex')
        field = data.get('field')
        value = data.get('value')
        
        if row_index is None or field is None or value is None:
            raise HTTPException(status_code=400, detail="Missing required parameters")
        
        # Get the validated data file
        validated_path = os.path.join("corrected", f"{file_id}_validated.json")
        if not os.path.exists(validated_path):
            raise HTTPException(status_code=404, detail="Validated data not found")
        
        with open(validated_path, 'r') as f:
            validated_data = json.load(f)
        
        # Update the cell
        if row_index < len(validated_data['data']):
            validated_data['data'][row_index][field] = value
            
            # Save the updated data
            with open(validated_path, 'w') as f:
                json.dump(validated_data, f, indent=2)
            
            return {
                "status": "success",
                "message": "Cell updated successfully"
            }
        else:
            raise HTTPException(status_code=400, detail="Invalid row index")
            
    except Exception as e:
        logger.error(f"Error updating cell: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Cell update failed: {str(e)}")

@app.get("/excel-data/{file_id}")
async def get_excel_data(file_id: str):
    """Get the converted Excel data for a file"""
    try:
        # Check for validated data first
        validated_path = os.path.join("corrected", f"{file_id}_validated.json")
        if os.path.exists(validated_path):
            with open(validated_path, 'r') as f:
                validated_data = json.load(f)
            return validated_data.get('data', [])
        
        # Check for converted Excel file
        excel_path = os.path.join("converted", f"{file_id}.xlsx")
        if os.path.exists(excel_path):
            df = pd.read_excel(excel_path)
            return df.to_dict('records')
        
        # Check for converted CSV file
        csv_path = os.path.join("converted", f"{file_id}.csv")
        if os.path.exists(csv_path):
            df = pd.read_csv(csv_path)
            return df.to_dict('records')
        
        raise HTTPException(status_code=404, detail="No converted data found")
        
    except Exception as e:
        logger.error(f"Error getting Excel data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get Excel data: {str(e)}")

@app.post("/gst/generate-reports/{file_id}")
async def generate_gst_reports(file_id: str, request: Request):
    """Generate GST reports for the converted data"""
    try:
        data = await request.json()
        period = data.get('period', datetime.now().strftime('%Y-%m'))
        
        # Get the validated data
        validated_path = os.path.join("corrected", f"{file_id}_validated.json")
        if not os.path.exists(validated_path):
            raise HTTPException(status_code=404, detail="Validated data not found")
        
        with open(validated_path, 'r') as f:
            validated_data = json.load(f)
        
        entries = validated_data.get('data', [])
        
        # Generate GSTR-1 data
        gstr1_data = []
        gstr3b_data = {
            "tax_period": period,
            "total_turnover": 0,
            "taxable_turnover": 0,
            "igst": 0,
            "cgst": 0,
            "sgst": 0,
            "cess": 0
        }
        
        total_turnover = 0
        taxable_turnover = 0
        
        for entry in entries:
            amount = float(entry.get('amount', 0))
            tax_rate = float(entry.get('taxRate', 0))
            gstin = entry.get('gstin', '')
            
            total_turnover += amount
            
            if tax_rate > 0:
                taxable_turnover += amount
                tax_amount = amount * (tax_rate / 100)
                
                # For simplicity, assume equal CGST and SGST for intra-state
                cgst_amount = tax_amount / 2
                sgst_amount = tax_amount / 2
                
                gstr3b_data["cgst"] += cgst_amount
                gstr3b_data["sgst"] += sgst_amount
                
                gstr1_data.append({
                    "gstin": gstin,
                    "invoice_number": entry.get('voucherNo', ''),
                    "invoice_date": entry.get('date', ''),
                    "invoice_value": amount,
                    "taxable_value": amount,
                    "tax_rate": tax_rate,
                    "cgst_amount": cgst_amount,
                    "sgst_amount": sgst_amount,
                    "igst_amount": 0,
                    "cess_amount": 0
                })
        
        gstr3b_data["total_turnover"] = total_turnover
        gstr3b_data["taxable_turnover"] = taxable_turnover
        
        # Save reports
        report_dir = "exports"
        os.makedirs(report_dir, exist_ok=True)
        
        gstr1_path = os.path.join(report_dir, f"GSTR1_{file_id}_{period}.json")
        gstr3b_path = os.path.join(report_dir, f"GSTR3B_{file_id}_{period}.json")
        
        with open(gstr1_path, 'w') as f:
            json.dump(gstr1_data, f, indent=2)
        
        with open(gstr3b_path, 'w') as f:
            json.dump(gstr3b_data, f, indent=2)
        
        return {
            "status": "success",
            "message": "GST reports generated successfully",
            "gstr1_file": os.path.basename(gstr1_path),
            "gstr3b_file": os.path.basename(gstr3b_path),
            "gstr1_data": gstr1_data,
            "gstr3b_data": gstr3b_data
        }
        
    except Exception as e:
        logger.error(f"Error generating GST reports: {str(e)}")
        raise HTTPException(status_code=500, detail=f"GST report generation failed: {str(e)}")

# Add after other initialization code, before routes
def cleanup_old_files():
    """Clean up old files to free up disk space"""
    current_time = time.time()
    max_age = 24 * 60 * 60  # 24 hours in seconds
    
    for directory in [UPLOAD_DIR, CONVERTED_DIR]:
        if not os.path.exists(directory):
            continue
            
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            if os.path.isfile(file_path):
                file_age = current_time - os.path.getmtime(file_path)
                if file_age > max_age:
                    try:
                        os.remove(file_path)
                        logger.info(f"Cleaned up old file: {file_path}")
                    except Exception as e:
                        logger.warning(f"Failed to clean up {file_path}: {str(e)}")

def memory_monitor():
    """Monitor memory usage and clean up cache if needed"""
    while True:
        try:
            memory_percent = psutil.virtual_memory().percent
            if memory_percent > 85:  # If memory usage > 85%
                logger.warning(f"High memory usage: {memory_percent}%")
                # Clear file cache
                file_cache.clear()
                logger.info("Cleared file cache due to high memory usage")
                
            time.sleep(300)  # Check every 5 minutes
        except Exception as e:
            logger.error(f"Error in memory monitor: {str(e)}")
            time.sleep(300)

def file_cleanup_scheduler():
    """Schedule file cleanup every hour"""
    while True:
        try:
            cleanup_old_files()
            time.sleep(3600)  # Run every hour
        except Exception as e:
            logger.error(f"Error in file cleanup: {str(e)}")
            time.sleep(3600)

# Start background threads for maintenance
cleanup_thread = threading.Thread(target=file_cleanup_scheduler, daemon=True)
cleanup_thread.start()

memory_thread = threading.Thread(target=memory_monitor, daemon=True)
memory_thread.start()

logger.info("Background maintenance threads started")

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True) 
    